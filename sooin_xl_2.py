#기본 설정
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from subprocess import CREATE_NO_WINDOW
serv=Service("C:/chromedriver/chromedriver.exe")
serv.creation_flags = CREATE_NO_WINDOW   
driver = webdriver.Chrome(service = serv)

#모듈 불러오기
import time, datetime, os # timesleep, 파일 이름을 위한 현재 시간, 파일 열기 모듈
import openpyxl # 파이썬이 엑셀 열게 해주는 모듈  

#파일 이름 저장
now=str(datetime.datetime.now()) [:16] #시간 불러오기-초 단위까지 자르기
folderName=now.replace(':','_') #:을 _로 바꿈
os.mkdir(folderName) #mkdir= make directory('폴더 이름')

#검색, 저장        
key_word=['비트코인','이더리움','세종시 맛집','살 빼는 운동'] #검색어

wb=openpyxl.Workbook() #워크북 열기

for i in range(len(key_word)): #key_word 리스트 길이만큼 반복
    ws=wb.create_sheet() #worksheet 생성
    ws.title=key_word[i] #worksheet 이름을 검색어로 변경
    ws.column_dimensions['A'].width=90 #A열 너비 조절
    
    url = "https://search.naver.com/search.naver?where=news&ie=utf8&sm=nws_hty&query=" + key_word[i] #링크
    driver.get(url) #링크 주소 열기
    time.sleep(2)
    
    list_news=driver.find_element('class name','list_news') #뉴스 찾기
    news_boxes=list_news.find_elements('class name', 'bx') #뉴스 저장

    for j in range(len(news_boxes)): #검색어 갯수만큼 반복
        driver.execute_script('arguments[0].scrollIntoView(true);', news_boxes[j]) #뉴스까지 스크롤
        file=f'{folderName}/{i+1}_{key_word[i]}-{j+1}.png' #파일 이름 저장
        news_boxes[j].screenshot(file) #뉴스 스크린샷

        ws.row_dimensions[j+1].height=100 #뉴스 행 높이 조절
        img=openpyxl.drawing.image.Image(file) #스크린샷을 엑셀 삽입가능 파일로 저장
        ws.add_image(img,f'A{j+1}') #셀에 이미지(스크린샷) 삽입


        title=news_boxes[j].find_element('class name','news_tit') #뉴스 타이틀 개체 변수 저장
        print(j+1, title.text) #셀에 프린트

        link=title.get_attribute('href') #개체 링크 저장
        ws[f'B{j+1}'].value='기사링크' 
        ws[f'B{j+1}'].hyperlink=link #B열에 링크 삽입

    print()

wb.remove(wb["Sheet"])  #필요없는 시트 삭제
wb.save(f'{folderName}/{folderName}_결과.xlsx') #엑셀 저장
